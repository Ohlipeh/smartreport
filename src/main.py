import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Caminhos
input_file = "data/vendas.csv"
output_file = "output/relatorio.xlsx"

# Ler dados
df = pd.read_csv(input_file)

# Criar coluna de faturamento
df["faturamento"] = df["quantidade"] * df["preco_unitario"]

# Resumo geral
total_vendas = df["faturamento"].sum()
media_vendas = df["faturamento"].mean()
produto_mais_vendido = df.groupby("produto")["quantidade"].sum().idxmax()

# Agrupar por categoria
resumo_categoria = df.groupby("categoria")["faturamento"].sum().reset_index()

with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Dados Brutos", index=False)
    resumo_categoria.to_excel(writer, sheet_name="Resumo por Categoria", index=False)

    resumo_executivo = pd.DataFrame(
        {
            "Indicador": ["Total de Vendas", "Média de Vendas", "Produto Mais Vendido"],
            "Valor": [total_vendas, media_vendas, produto_mais_vendido],
        }
    )

    resumo_executivo.to_excel(writer, sheet_name="Resumo Executivo", index=False)

    workbook = writer.book

    # Formatar valores monetários
    for sheet_name in ["Dados Brutos", "Resumo por Categoria"]:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "R$ #,##0.00"

        # Ajustar largura automática
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[column_letter].width = max_length + 2

    # Destacar título do resumo executivo
    resumo_sheet = workbook["Resumo Executivo"]
    resumo_sheet["A1"].font = Font(bold=True)
    resumo_sheet["B1"].font = Font(bold=True)

print("Relatório profissional gerado com sucesso!")
